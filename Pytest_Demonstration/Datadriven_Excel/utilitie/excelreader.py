import openpyxl

def getdata(path,sheet):
    res = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet]
    trow = sheet.max_row
    tcol = sheet.max_column

    for i in range(2,trow + 1):
        rlist = []
        for j in range(1,tcol + 1):
            rlist.append(sheet.cell(i,j).value)
        res.append(rlist)
    
    return res


